import xlrd,xlwt
from openpyxl import load_workbook
from openpyxl import Workbook

#workbook = xlrd.open_workbook('数据.xls')
workbook= load_workbook("数据.xlsx")
sheet1 = workbook["Sheet1"]
cols = sheet1["A"]

    
u=list()

for i in cols:
    if(i.value not in ['','------------------------------'] and i.value != None):
        u.append(i.value)


wb = Workbook()
sheet1 = wb.active

for i in range(len(u)):
    sheet1["A%d" % (i+1)].value=u[i]

y=0
p=list()
l=list()                  # l为等开的 位置
for i in range(len(u)):
    if(u[i].split(' ')[-1]=='等开'):
        l.append(i)
        
for i in range(len(u)):
    if(i in l):
        y=i
        p.append("")
    else:
        if(int(u[i].split('-')[0])%8 in [2,4,5,7]):
            p.append('中')
        else:
            p.append('挂')
        '''if((i-1-y)%3==1 and y!=0):
            p.append('挂')
        elif((i-y)%3==1 and y==0):
            p.append('挂')
        else:
            p.append("中")'''
        
for i in range(len(p)):
    sheet1["B%d" % (i+1)].value=p[i]




multi=1
d=dict()
n=0
multi_f=1
for i in range(len(p)):
    multi=multi_f
    if(u[i].split(' ')[-1]==p[i]):
        multi=1
        d[i]=multi
    elif(u[i].split(' ')[-1]=='等开'):
        d[i]=''
    else:
        d[i]=multi
        sheet1["C%d" % (i+1)].value='错'
        multi=multi*2
            

wb.save('1.xlsx')

