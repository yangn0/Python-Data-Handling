import openpyxl

NUM=20

excel=openpyxl.load_workbook('ETF(1).xlsx')
sheets=excel.sheetnames

for sheet in sheets:
    print(sheet)
    ws=excel[sheet]
    #row_date=ws["A"]
    row_open=ws["E"]
    row_close=ws['H']
    row_turn=ws['M']
    u=0
    turn_list=[]
    turnavg_num=0
    for i in range(4,len(row_open)):
        u+=1
        try:        
            turn_list.append(row_turn[i].value)
        except IndexError:
            break
        if(u==1):
            u_open=row_open[i].value
        elif(u==NUM):
            u_close=row_close[i].value
            oc=float(u_close)/float(u_open)
            ws["N%s"%(i+1)]=oc
            turn_avg=sum(turn_list)/len(turn_list)
            ws["O%s"%(i+1)]=turn_avg
            turnavg_num+=1
            if(turnavg_num%2==0):
                ws["P%s"%(i+1)]=turn_avg/turn_avg_last
            turn_avg_last=turn_avg
            turn_list=[]
            u=0
excel.save('结果.xlsx')
        
        
    
        