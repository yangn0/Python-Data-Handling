import openpyxl

print('请确保 10年确权数据.xlsm 和 承包方调查信息表.xlsx，与本软件在同一目录下。')
input('准备好后 摁回车 开始汇总……')
path='10年确权数据.xlsm'
#path=input('请输入 确权数据表 名称（需带后缀）：')
path1='承包方调查信息表.xlsx'
try:
    excel1=openpyxl.load_workbook(path1)
except:
    print('未找到'+path1)
    a=input()
print(path1+'导入成功')

try:
    excel=openpyxl.load_workbook(path,read_only=True)
except:
    print('未找到'+path)
    a=input()
print(path+'导入成功')

print('正在汇总 请稍等……')
max_row=excel['Sheet1'].max_row
d=dict()
for i in range(1,max_row):
    if(excel['Sheet1'].cell(row=i,column=1).value == '重庆市璧山县农村土地承包经营权证登记簿'):
        name=excel['Sheet1'].cell(row=i+3,column=5).value#承包方代表姓名

        '''human_num=excel['Sheet1'].cell(row=i+3,column=13).value#身份证号
        if(human_num==None):
            human_num='' '''

        human_num=''

        book_num=excel['Sheet1'].cell(row=i+2,column=13).value#证书号
        #二轮合同面积
        two_area=list()
        u=0
        while(1):
            a=excel['Sheet1'].cell(row=i+u+12,column=7).value
            if(a==None):
                break           
            two_area.append(a)
            u+=1
        #2010年确权面积亩
        confirm_area=list()
        u=0
        while(1):
            a=excel['Sheet1'].cell(row=i+u+12,column=8).value
            if(a==None):
                break           
            confirm_area.append(a)
            u+=1
        #总面积
        area=excel['Sheet1'].cell(row=i+9,column=1).value
        a=list()
        for u in area:
            if u>='0' and  u<='9' or u=='.':
                a+=u
        area=''.join(a)
        l=[name,human_num,book_num,two_area,confirm_area,area]
        d[name]=l

for i in range(excel1['鱼鳞图'].max_row):
    cell=excel1['鱼鳞图'].cell(row=6+i,column=2).value
    if(cell!=None):
        # try:
        #     excel1['鱼鳞图'].cell(row=6+i,column=8).value=d[cell][1]
        # except:
        #     continue
        try:
            u=0
            for u in range(len(d[cell][3])):
                if(excel1['鱼鳞图'].cell(row=6+i+u,column=2).value==None or u==0):
                    excel1['鱼鳞图'].cell(row=6+i+u,column=24).value=d[cell][3][u]
                else:
                    break
            u=0
            for u in range(len(d[cell][4])):
                if(excel1['鱼鳞图'].cell(row=6+i+u,column=2).value==None or u==0):
                    excel1['鱼鳞图'].cell(row=6+i+u,column=43).value=d[cell][4][u]
                else:
                    break
            excel1['鱼鳞图'].cell(row=6+i,column=45).value=d[cell][2]
            excel1['鱼鳞图'].cell(row=6+i,column=46).value=d[cell][5]
        except:
            continue
excel1.save('承包地块调查表汇总.xlsx')
print('汇总完成。已保存为 承包地块调查表汇总.xlsx。')
a=input()
