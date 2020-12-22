from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import  PatternFill
#d = ws1.cell(row=4, column=2, value=10)

print("统计 5个位置的 大小 或 单双")
name=input("文件位置：")
#name="澳洲幸运5.xlsx"

mode=int(input("选择模式 1.大小 2.单双:"))

time=int(input("周期："))

time_l=input("大 或者 单 的位置：").split(' ')

time_l=list(map(int,time_l))

print("正在计算请稍等……")

workbook= load_workbook(name)
sheet1 = workbook["Sheet1"]

position1 = sheet1["C"]
position2 = sheet1["D"]
position3 = sheet1["E"]
position4 = sheet1["F"]
position5 = sheet1["G"]
position=[position1,position2,position3,position4,position5]
p_d={1:[],2:[],3:[],4:[],5:[]}

for u in range(1,6):
    for i in position[u-1][1:]:
        p_d[u].append(int(i.value))
    
wb = Workbook()
sheet2 = wb.active

n=1
for i in sheet1["A"][1:]:
    sheet2.cell(row=n, column=1, value=i.value)
    n+=1

n=1
for i in sheet1["B"][1:]:
    sheet2.cell(row=n, column=2, value=int(i.value))
    n+=1

for u in range(1,6):
    n=1
    for i in p_d[u]:
        sheet2.cell(row=n,column=u*6-3,value=i)
        n+=1

if(mode==1):
    for u in range(1,6):
        n=1
        for i in p_d[u]:
            if(sheet2.cell(row=n,column=u*6-3).value in [5,6,7,8,9]):
                sheet2.cell(row=n,column=u*6-3+1,value="大")
            else:
                sheet2.cell(row=n,column=u*6-3+1,value="小")
            n+=1

    for u in range(1,6):
        n=1
        for i in p_d[u]:
            if(int(str(sheet2.cell(row=n,column=2).value)[-2:]) %time in time_l):
                sheet2.cell(row=n,column=u*6-3+2,value="大")
            else:
                sheet2.cell(row=n,column=u*6-3+2,value="小")
            n+=1
else:
    for u in range(1,6):
        n=1
        for i in p_d[u]:
            if(sheet2.cell(row=n,column=u*6-3).value %2 ==1 ):
                sheet2.cell(row=n,column=u*6-3+1,value="单")
            else:
                sheet2.cell(row=n,column=u*6-3+1,value="双")
            n+=1

    for u in range(1,6):
        n=1
        for i in p_d[u]:
            if(int(str(sheet2.cell(row=n,column=2).value)[-2:]) %time in time_l):
                sheet2.cell(row=n,column=u*6-3+2,value="单")
            else:
                sheet2.cell(row=n,column=u*6-3+2,value="双")
            n+=1

for u in range(1,6):
    n=1
    for i in p_d[u]:
        if(sheet2.cell(row=n,column=u*6-3+1).value == sheet2.cell(row=n,column=u*6-3+2).value):
            sheet2.cell(row=n,column=u*6-3+3,value="对")
        else:
            sheet2.cell(row=n,column=u*6-3+3,value="错")
        n+=1
        
true_or_false={1:[],2:[],3:[],4:[],5:[]}
for u in range(1,6):
    n=1
    for i in p_d[u]:
        if(sheet2.cell(row=n,column=u*6-3+1).value == sheet2.cell(row=n,column=u*6-3+2).value):
            sheet2.cell(row=n,column=u*6-3+3,value="对")
            true_or_false[u].append("对")
        else:
            sheet2.cell(row=n,column=u*6-3+3,value="错")
            true_or_false[u].append("错")
        n+=1

for u in range(1,6):
    multi=1
    n=1
    for i in true_or_false[u]:
        if (i=="对"):
            sheet2.cell(row=n,column=u*6-3+4,value=multi)
            multi=1
        else:
            sheet2.cell(row=n,column=u*6-3+4,value=multi)
            multi=multi*2
        n+=1
        
fill = PatternFill("solid", fgColor="1874CD")
for u in range(1,6):
    multi=1
    n=1
    for i in true_or_false[u]:
        if(multi==0):
            if(int(str(sheet2.cell(row=n,column=2).value)[-2:]) %time == 1):
                multi=1
            else:
                n+=1
                sheet2.cell(row=n,column=u*6-3+5,value=multi)
                continue
            
        if (i=="对"):
            st=sheet2.cell(row=n,column=u*6-3+5,value=multi)
            multi=0
        else:
            st=sheet2.cell(row=n,column=u*6-3+5,value=multi)
            multi=multi*2
        if(int(str(sheet2.cell(row=n,column=2).value)[-2:]) %time == 1):
            sheet2.cell(row=n,column=u*6-3+5).fill=fill
        n+=1

if(mode==1):
    wb.save("大小.xlsx")
else:
    wb.save("单双.xlsx")

