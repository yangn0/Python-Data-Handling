from openpyxl import load_workbook
from openpyxl import Workbook
#d = ws1.cell(row=4, column=2, value=10)
name="新疆SSC.xlsx"
workbook= load_workbook(name)
sheet1 = workbook["Sheet1"]

position1 = sheet1["C"]
position2 = sheet1["D"]
position3 = sheet1["E"]
position4 = sheet1["F"]
position5 = sheet1["G"]
position=[position1,position2,position3,position4,position5]
p_d={1:[],2:[],3:[],4:[],5:[]}

for u in range(1,6):
    for i in position[u-1][1:]:
        p_d[u].append(int(i.value))
    
wb = Workbook()
sheet1 = wb.active

def push_in(p1,col):
    d=dict()
    d1=dict()
    while(1):
        if(len(p1)==0):
            break
        if(p1.pop() in [5,6,7,8,9]):
            n=1
            try:
                while(p1.pop()in [5,6,7,8,9]):
                    n+=1
            except:
                break
            if(n in d.keys()):
                d[n]+=1
            else:
                d[n]=1
        else:
            n=1
            try:
                while(p1.pop() in [0,1,2,3,4]):
                    n+=1
            except:
                break
            if(n in d1.keys()):
                d1[n]+=1
            else:
                d1[n]=1
    for i in d.keys():
        sheet1.cell(row=i, column=col, value=d[i])
    for i in d1.keys():
        sheet1.cell(row=i, column=col+1, value=d1[i])
for u in range(1,6):
    push_in(p_d[u],u*3-1)
    
wb.save("大小_"+name)

