from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import  PatternFill
#d = ws1.cell(row=4, column=2, value=10)
name="幸运飞艇.xlsx"
workbook= load_workbook(name)
sheet1 = workbook["Sheet1"]

position1 = sheet1["C"]
position2 = sheet1["D"]
position3 = sheet1["E"]
position4 = sheet1["F"]
position5 = sheet1["G"]
position6 = sheet1["H"]
position7 = sheet1["I"]
position8 = sheet1["J"]
position9 = sheet1["K"]
position10= sheet1["L"]
position=[position1,position2,position3,position4,position5,position6,position7,position8,position9,position10]
p_d={1:[],2:[],3:[],4:[],5:[],6:[],7:[],8:[],9:[],10:[]}

for u in range(1,11):
    for i in position[u-1][1:]:
        p_d[u].append(int(i.value))
    
wb = Workbook()
sheet2 = wb.active

n=1
for i in sheet1["A"][1:]:
    sheet2.cell(row=n, column=1, value=i.value)
    n+=1

n=1
for i in sheet1["B"][1:]:
    sheet2.cell(row=n, column=2, value=int(i.value))
    n+=1

for u in range(1,11):
    n=1
    for i in p_d[u]:
        sheet2.cell(row=n,column=u*6-3,value=i)
        n+=1

for u in range(1,11):
    n=1
    for i in p_d[u]:
        if(sheet2.cell(row=n,column=u*6-3).value in [5,6,7,8,9]):
            sheet2.cell(row=n,column=u*6-3+1,value="大")
        else:
            sheet2.cell(row=n,column=u*6-3+1,value="小")
        n+=1

for u in range(1,11):
    n=1
    for i in p_d[u]:
        if(int(str(sheet2.cell(row=n,column=2).value)[-2:]) %8 in [1,3,6,0]):
            sheet2.cell(row=n,column=u*6-3+2,value="大")
        else:
            sheet2.cell(row=n,column=u*6-3+2,value="小")
        n+=1

for u in range(1,11):
    n=1
    for i in p_d[u]:
        if(sheet2.cell(row=n,column=u*6-3+1).value == sheet2.cell(row=n,column=u*6-3+2).value):
            sheet2.cell(row=n,column=u*6-3+3,value="对")
        else:
            sheet2.cell(row=n,column=u*6-3+3,value="错")
        n+=1
        
true_or_false={1:[],2:[],3:[],4:[],5:[],6:[],7:[],8:[],9:[],10:[]}
for u in range(1,11):
    n=1
    for i in p_d[u]:
        if(sheet2.cell(row=n,column=u*6-3+1).value == sheet2.cell(row=n,column=u*6-3+2).value):
            sheet2.cell(row=n,column=u*6-3+3,value="对")
            true_or_false[u].append("对")
        else:
            sheet2.cell(row=n,column=u*6-3+3,value="错")
            true_or_false[u].append("错")
        n+=1

for u in range(1,11):
    multi=1
    n=1
    for i in true_or_false[u]:
        if (i=="对"):
            sheet2.cell(row=n,column=u*6-3+4,value=multi)
            multi=1
        else:
            sheet2.cell(row=n,column=u*6-3+4,value=multi)
            multi=multi*2
        n+=1
        
fill = PatternFill("solid", fgColor="1874CD")
for u in range(1,11):
    multi=1
    n=1
    c=0
    for i in true_or_false[u]:
        if(c!=0):
            c-=1
            continue
        if (i=="对"):
            st=sheet2.cell(row=n,column=u*6-3+5,value=multi)
            multi=1
        else:
            st=sheet2.cell(row=n,column=u*6-3+5,value=multi)
            multi=multi*2
        if(multi==32):
            n+=6
            c=6
        st.fill=fill
        n+=1

   
wb.save("大小_"+name)

