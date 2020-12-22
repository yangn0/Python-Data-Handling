import openpyxl
from openpyxl import load_workbook

input("确保 总面积.xlsx和需分摊面积.xlsx 与本程序在同一目录。按回车继续……")
print("正在操作请稍等.")
wb1 = load_workbook('总面积.xlsx')
ws = wb1.active
data= ws.rows
d=dict()
for i in data:
    d[i[0].value]=i[1].value

wb2 = load_workbook('需分摊面积.xlsx')
ws = wb2.active
data= ws.rows

d_num=dict()
for i in data:
    if i[0].value not in d_num:
        d_num[i[0].value]=1
    else:
        d_num[i[0].value]+=1

data= ws.rows
for i in data:
    try:
        i[23].value=d[i[0].value]/d_num[i[0].value]
    except:
        pass
    

wb2.save('结果.xlsx')

input("已完成。按回车退出……")